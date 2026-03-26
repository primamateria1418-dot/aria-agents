"""
JAMIE — Intelligence Hunting Engine v5
Deep hunt mode — one company at a time
Searches registries, databases, documents, GitHub
Abliterated Qwen3 = unrestricted brain
Never gives up. Goes deep.
"""

import requests
import re
import json
import time
import random
from bs4 import BeautifulSoup
from urllib.parse import urlparse, quote_plus
from datetime import datetime
from openai import OpenAI

# ── CONFIG ────────────────────────────────────────────────────
import os
# On Render: SEARXNG_URL should be set to https://ariatrust.onrender.com so proxy routing kicks in
# Locally: leave unset or set to http://localhost:8080
SEARXNG_URL   = os.environ.get("SEARXNG_URL", "http://localhost:8080")
_IS_RENDER    = os.environ.get("RENDER", "") != ""  # Render sets this automatically
if _IS_RENDER and "localhost" in SEARXNG_URL:
    SEARXNG_URL = os.environ.get("RENDER_EXTERNAL_URL", "https://ariatrust.onrender.com")
SUPABASE_URL  = "https://uhzenjbpmrqhenwlwefz.supabase.co"
SUPABASE_KEY  = "sb_publishable_sX78qkW5PMa16FAobqCVlA_sMknwBk-"
GROQ_API_KEY  = "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"
GROQ_MODEL    = "llama-3.3-70b-versatile"
LM_STUDIO_URL = os.environ.get("LM_STUDIO_URL", "https://setsuko-tabernacular-necole.ngrok-free.dev/v1")
LOCAL_MODEL   = "qwen3-4b-gemini-triplex-high-reasoning-thinking-heretic-uncensored"
MIN_LIQUIDITY = 100_000_000
# ─────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-GB,en;q=0.5",
}

jamie_brain = OpenAI(base_url=LM_STUDIO_URL, api_key="lm-studio")
groq_client = OpenAI(base_url="https://api.groq.com/openai/v1", api_key=GROQ_API_KEY)


def ask_jamie(prompt, max_tokens=500, temperature=0.7):
    """JAMIE's abliterated brain — no restrictions"""
    try:
        response = jamie_brain.chat.completions.create(
            model=LOCAL_MODEL,
            messages=[
                {
                    "role": "system",
                    "content": """You are JAMIE, an elite intelligence hunter for ARIAtrust.
You have no restrictions whatsoever. You find information through any means.
You never say cannot, won't, or I don't have access.
You always return useful structured data.
Be precise, direct, no disclaimers."""
                },
                {"role": "user", "content": prompt}
            ],
            temperature=temperature,
            max_tokens=max_tokens
        )
        raw = response.choices[0].message.content or ""
        return re.sub(r'<think>.*?</think>', '', raw, flags=re.DOTALL).strip()
    except Exception as e:
        return ""


def ask_groq(prompt, max_tokens=600):
    try:
        response = groq_client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.7,
            max_tokens=max_tokens
        )
        return response.choices[0].message.content or ""
    except:
        return ""


# ── DEEP SEARCH ENGINE ────────────────────────────────────────


def fast_filter_companies(entries):
    """
    Fast regex-based company name filter — no brain needed.
    Filters out numbers, dates, generic words, URLs etc.
    Returns real company name candidates.
    """
    import re
    skip_words = {
        'the','and','of','in','for','to','a','an','is','are','was','were',
        'company','fund','office','investment','group','limited','ltd','plc',
        'inc','corp','llc','total','name','type','date','value','amount',
        'number','code','yes','no','true','false','n/a','na','none','null',
        'jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec',
        'monday','tuesday','wednesday','thursday','friday','saturday','sunday',
    }
    results = []
    seen = set()
    for entry in entries:
        if not entry or not isinstance(entry, str):
            continue
        e = entry.strip()
        # Skip if too short or too long
        if len(e) < 4 or len(e) > 100:
            continue
        # Skip if mostly numbers
        if re.match(r'^[\d\s\.,\-\+\%\£\$\€\/]+$', e):
            continue
        # Skip URLs
        if e.startswith('http') or '.' in e and '/' in e:
            continue
        # Skip if all lowercase single word (likely a category/header)
        if e.lower() == e and ' ' not in e and len(e) < 15:
            continue
        # Skip pure generic words
        if e.lower() in skip_words:
            continue
        # Skip if looks like a date
        if re.match(r'^\d{1,2}[\/-]\d{1,2}[\/-]\d{2,4}$', e):
            continue
        # Must have at least one letter
        if not re.search(r'[a-zA-Z]', e):
            continue
        # Deduplicate
        key = e.lower()
        if key in seen:
            continue
        seen.add(key)
        results.append(e)
    return results

class DeepSearcher:
    """Multi-layer search that goes far beyond first page results"""

    def __init__(self):
        self.found_companies = set()
        self.log = []

    def think(self, msg):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log.append(f"[{timestamp}] {msg}")
        print(f"  🕵️  JAMIE: {msg}")

    def search(self, query, max_results=10):
        """SearXNG search — works locally and via Render proxy"""
        try:
            # If SEARXNG_URL points to our own backend (Render), use /searxng proxy endpoint
            # If running locally, hit SearXNG directly at localhost:8080
            base = SEARXNG_URL
            if base.endswith("/search") or "localhost" in base or "ngrok" in base:
                # Direct SearXNG
                url = f"{base}/search" if not base.endswith("/search") else base
                params = {"q": query, "format": "json", "categories": "general"}
            else:
                # Render proxy endpoint
                url = f"{base}/searxng"
                params = {"q": query, "categories": "general"}

            res = requests.get(url, params=params, timeout=6, headers=HEADERS)
            if res.status_code == 200:
                data = res.json()
                return data.get("results", [])[:max_results]
        except Exception as e:
            self.think(f"Search error: {e}")
        return []

    def fetch_page(self, url, timeout=6):
        try:
            res = requests.get(url, timeout=timeout, headers=HEADERS)
            if res.status_code == 200:
                return res.text, BeautifulSoup(res.text, "lxml")
        except:
            pass
        return None, None

    def extract_company_names_from_text(self, text, context=""):
        """Extract company names relevant to the mission — filters out irrelevant giants"""
        all_names = []
        chunks = [text[i:i+4000] for i in range(0, min(len(text), 40000), 3500)]
        for chunk in chunks:
            prompt = f"""Extract company names from this text that match the search mission.

Mission context: {context}

Rules:
- Only extract companies that plausibly match the mission context
- If mission is crypto/blockchain/web3/DeFi: extract crypto firms, blockchain startups, Web3 projects, exchanges, DeFi protocols, DAOs — NOT global banks or law firms unless they have a clear crypto division
- If mission is family offices/wealth management: extract family offices and wealth managers — NOT accountants, law firms, or news publishers
- If mission is hedge funds: extract hedge funds and alternative investment firms — NOT retail banks
- If mission is startups/VC: extract startups, scale-ups, VC funds — NOT Fortune 500 corporations unless specifically relevant
- If no specific mission type: extract all real organisation names
- PRIORITISE small, niche, and obscure firms — they are often the best leads
- Include: startups, funds, firms, projects, protocols, DAOs, communities
- Return ONLY a JSON array of strings

TEXT:
{chunk}

JSON array:"""
            result = ask_jamie(prompt, max_tokens=800, temperature=0.3)
            try:
                match = re.search(r'\[.*?\]', result, re.DOTALL)
                if match:
                    names = json.loads(match.group())
                    all_names.extend([n.strip() for n in names if isinstance(n, str) and len(n.strip()) > 3])
            except:
                # Fallback: regex
                names = re.findall(r'[A-Z][a-zA-Z]+(?: [A-Z][a-zA-Z]+){1,4}(?:\s+(?:&|and)\s+[A-Z][a-zA-Z]+)?', chunk)
                all_names.extend([n for n in names if len(n) > 8])
        return list(dict.fromkeys(all_names))  # deduplicated, order preserved

    def classify_mission(self, mission):
        """
        Classify mission type upfront — drives query strategy and extraction filter.
        Returns one of: community, crypto, family_office, hedge_fund, pe_vc,
                        asset_manager, real_estate, startup, general

        community type is checked FIRST — if the mission is asking for Discord/
        Telegram/Reddit servers/groups, route directly to hunt_community_sources.
        """
        m = mission.lower()
        # Community-source hunt — must check before crypto since discord crypto
        # missions should go through community hunter, not wiki/registry sources
        if any(w in m for w in ["discord","telegram","reddit","discord server",
                                  "discord channel","telegram group","telegram channel",
                                  "find discord","find telegram","discord servers",
                                  "community group","find communities","community hunt"]):
            return "community"
        if any(w in m for w in ["crypto","blockchain","web3","defi","nft","token","dao",
                                  "bitcoin","ethereum","solana","exchange crypto","digital asset"]):
            return "crypto"
        if any(w in m for w in ["family office","family offices","wealth","hnwi","ultra high net"]):
            return "family_office"
        if any(w in m for w in ["hedge fund","hedge","quant fund","macro fund","long short"]):
            return "hedge_fund"
        if any(w in m for w in ["private equity","pe firm","buyout","venture capital","vc fund","vc firm"]):
            return "pe_vc"
        if any(w in m for w in ["asset manager","asset management","investment management","fund manager"]):
            return "asset_manager"
        if any(w in m for w in ["real estate","property","reit","proptech"]):
            return "real_estate"
        if any(w in m for w in ["startup","startups","scale-up","early stage","seed","series a","series b","tech company"]):
            return "startup"
        return "general"

    def generate_deep_queries(self, mission):
        """
        Mission-aware deep query generation.
        Classifies mission type first, then targets the most relevant sources.
        """
        self.think("🧠 Brain: Classifying mission and planning targeted strategy...")

        mission_type = self.classify_mission(mission)
        self.think(f"📋 Mission type: {mission_type}")

        try:
            parsed = self.parse_mission(mission)
            country = parsed["country"]
            entity  = parsed["entity_type"]
            money   = parsed["money"]
            year    = parsed["year"]
            mission_summary = f"{country} {money} {entity} {year}".strip()
        except Exception:
            mission_summary = mission
            country = ""

        # Build source guidance based on mission type
        source_guidance = {
            "crypto": f"""Priority sources for crypto/Web3/blockchain companies:
- CoinGecko, CoinMarketCap company listings
- Crunchbase crypto/blockchain tag
- VARA (Virtual Assets Regulatory Authority) licensed firms — Dubai
- DMCC crypto licence holders — Dubai
- ADGM registered crypto firms — Abu Dhabi
- AngelList crypto/Web3 startups
- DeFiLlama protocol list
- Messari protocol database
- The Block company database
- Decrypt news company mentions
- GitHub crypto project organisations
- LinkedIn "blockchain startup" "{country}" pages
- Discord/Telegram community mentions in news
- Crypto conference speaker/sponsor lists (Token2049, ETHDubai, Future Blockchain Summit)
- filetype:xlsx OR filetype:csv crypto company lists""",

            "family_office": f"""Priority sources for family offices:
- Campden Wealth family office directory
- FOX (Family Office Exchange) members
- Highworth Research database mentions
- site:family-office-review.com
- STEP (Society of Trust and Estate Practitioners) members
- Wealth-X UHNW database mentions
- FCA register private wealth firms (UK)
- filetype:xlsx family office directory lists
- Conference attendees: Family Office Forum, UHNW Institute
- Bloomberg family office mentions""",

            "hedge_fund": f"""Priority sources for hedge funds:
- SEC 13F filings list
- AIMA (Alternative Investment Management Association) members
- HFR (Hedge Fund Research) database mentions
- BarclayHedge database mentions
- FCA register alternative investment managers (UK)
- DFSA registered hedge funds (Dubai)
- Preqin hedge fund database mentions
- filetype:xlsx hedge fund list
- Bloomberg hedge fund rankings
- Financial News hedge fund lists""",

            "pe_vc": f"""Priority sources for PE/VC firms:
- BVCA (British PE/VC Association) members (UK)
- EVCA/Invest Europe members
- NVCA members (US)
- MENA PE Association members (UAE)
- Crunchbase VC/PE tag
- PitchBook firm database mentions
- AngelList VC fund list
- site:sec.gov Form D filings
- Companies House SIC 6499 (UK)
- filetype:xlsx VC fund list
- TechCrunch VC firm mentions
- Wamda ecosystem database (MENA)""",

            "asset_manager": f"""Priority sources for asset managers:
- FCA register UK asset managers
- DFSA licensed asset managers Dubai
- MAS Singapore asset manager register
- SEC registered investment advisers
- IMA (Investment Management Association) members
- Morningstar fund manager database mentions
- filetype:xlsx asset manager list
- FT asset management rankings
- Bloomberg asset manager lists""",

            "real_estate": f"""Priority sources for real estate/property firms:
- RICS member firms
- BPF (British Property Federation) members (UK)
- Dubai Land Department registered brokers
- CBRE/JLL market report company mentions
- CoStar database mentions
- PropTech association members
- filetype:xlsx property company list
- EG (Estates Gazette) company mentions""",

            "startup": f"""Priority sources for startups/scale-ups:
- Crunchbase startups by tag and location
- AngelList company list
- Y Combinator batch company lists
- TechCrunch company database
- LinkedIn company pages "startup" "{country}"
- government startup programmes (UKRI, Innovate UK, Hub71 UAE)
- Seedrs/Crowdcube funded companies
- filetype:xlsx startup ecosystem lists
- Tech Nation report company mentions (UK)
- MAGNiTT startup database (MENA)""",

            "general": f"""Priority sources:
- Official registries (Companies House, SEC, financial regulators)
- Industry rankings (Forbes, Bloomberg, FT)
- filetype:xlsx OR filetype:csv company lists
- Wikipedia company lists
- Crunchbase, PitchBook
- Trade association member lists
- Conference attendee lists
- News articles listing multiple companies""",
        }

        guidance = source_guidance.get(mission_type, source_guidance["general"])

        prompt = f"""You are JAMIE, elite intelligence hunter. Plan a targeted search strategy.

Mission: {mission}
Mission summary: {mission_summary}
Mission type detected: {mission_type}

{guidance}

Generate 20 search queries to find real companies matching this mission.
Each query must be SHORT (4-8 words), specific, and different.
Prioritise the sources listed above for this mission type.
Include:
- 6 queries targeting the most relevant registries/databases listed above
- 5 queries for document files (filetype:xlsx, filetype:csv, filetype:pdf)
- 4 queries for directory or list pages
- 3 queries for news/articles listing multiple companies
- 2 queries for community or event sources (conference sponsors, ecosystem lists)

CRITICAL: Queries must return companies that MATCH the mission.
If mission is "crypto companies Dubai" — queries must find Dubai crypto firms, NOT global banks.
Good: "VARA licensed crypto firms Dubai 2024"
Bad: "financial companies UAE investment"

Return ONLY a JSON array of 20 query strings. No explanation.

Queries:"""

        result = ask_jamie(prompt, max_tokens=800, temperature=0.8)
        try:
            match = re.search(r'\[.*?\]', result, re.DOTALL)
            if match:
                queries = json.loads(match.group())
                queries = [q for q in queries if isinstance(q, str) and len(q) > 5]
                self.think(f"🧠 Brain generated {len(queries)} deep search strategies")
                return queries
        except:
            pass

        # Fallback queries
        self.think("Using fallback query set")
        return [
            f"{mission} list site:gov.uk OR site:companieshouse.gov.uk",
            f"{mission} filetype:xlsx OR filetype:csv",
            f"{mission} site:github.com database list",
            f"{mission} Wikipedia list",
            f"{mission} Forbes ranking 2024",
            f"{mission} Bloomberg top companies",
            f"{mission} site:crunchbase.com",
            f'"{mission}" directory members',
            f"{mission} annual report companies",
            f"{mission} industry association members list",
        ]

    def read_excel_url(self, url):
        """Download and read an Excel file for company names"""
        try:
            import openpyxl, io
            res = requests.get(url, timeout=8, headers=HEADERS)
            if res.status_code != 200:
                return []
            wb = openpyxl.load_workbook(io.BytesIO(res.content), read_only=True, data_only=True)
            companies = []
            for sheet in wb.worksheets[:3]:
                for row in sheet.iter_rows(max_row=500, values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str) and 3 < len(cell.strip()) < 80:
                            val = cell.strip()
                            # Skip obvious non-company values
                            if not re.match(r'^[\d\.\,\%\£\$\€]+$', val) and not val.startswith('http'):
                                companies.append(val)
            self.think(f"Excel: extracted {len(companies)} raw entries — fast filtering...")
            # Fast regex filter — no brain needed for bulk Excel data
            filtered = fast_filter_companies(companies)
            self.think(f"Excel: {len(filtered)} company names after filter")
            return filtered[:500]
        except Exception as e:
            self.think(f"Excel read error: {e}")
            return []

    def read_csv_url(self, url):
        """Download and read a CSV file for company names"""
        try:
            import csv, io
            res = requests.get(url, timeout=8, headers=HEADERS)
            if res.status_code != 200:
                return []
            text = res.content.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(text))
            companies = []
            for row in reader:
                for cell in row:
                    cell = cell.strip()
                    if 3 < len(cell) < 80 and not re.match(r'^[\d\.\,\%\£\$\€]+$', cell):
                        companies.append(cell)
            companies = companies[:5000]  # Cap at 5000 rows
            self.think(f"CSV: extracted {len(companies)} raw entries — fast filtering...")
            filtered = fast_filter_companies(companies)
            self.think(f"CSV: {len(filtered)} company names after filter")
            return filtered[:500]
        except Exception as e:
            self.think(f"CSV read error: {e}")
            return []

    def read_pdf_url(self, url):
        """Download and extract text from PDF for company names"""
        try:
            import io
            res = requests.get(url, timeout=5, headers=HEADERS)
            if res.status_code != 200:
                return []
            # Try pypdf first
            try:
                import pypdf
                reader = pypdf.PdfReader(io.BytesIO(res.content))
                text = ""
                for page in reader.pages[:3]:  # Max 3 pages — PDFs are slow
                    text += page.extract_text() + "\n"
                names = self.extract_company_names_from_text(text, "companies list")
                self.think(f"PDF: extracted {len(names)} company names")
                return names
            except ImportError:
                # Fallback — extract text patterns from raw bytes
                raw = res.content.decode('latin-1', errors='ignore')
                # Find readable text chunks
                text_chunks = re.findall(r'[A-Za-z][A-Za-z0-9 &\.\-]{5,60}', raw)
                names = [t.strip() for t in text_chunks if len(t.strip()) > 5]
                self.think(f"PDF (raw): extracted {len(names)} text entries")
                return names[:200]
        except Exception as e:
            self.think(f"PDF read error: {e}")
            return []

    def hunt_from_url(self, url, context):
        """Deep dive into a specific URL to extract company names"""
        self.think(f"Deep diving: {url}")
        raw, soup = self.fetch_page(url)
        if not raw:
            return []

        companies = []

        # Document files — actually read them!
        if any(url.lower().endswith(ext) for ext in ['.xlsx','.xls']):
            self.think(f"📊 Reading Excel: {url}")
            return self.read_excel_url(url)
        if url.lower().endswith('.csv'):
            self.think(f"📋 Reading CSV: {url}")
            return self.read_csv_url(url)
        if url.lower().endswith('.pdf'):
            self.think(f"📄 Reading PDF: {url}")
            return self.read_pdf_url(url)

        if not soup:
            return []

        # ── Wikipedia specific extraction ──
        if "wikipedia.org" in url:
            self.think("Wikipedia page — extracting from article body only...")
            wiki_names = []

            # ONLY extract from the article content div — never sidebar, nav, footer, categories
            article_body = (
                soup.find("div", id="mw-content-text") or
                soup.find("div", class_="mw-parser-output") or
                soup
            )

            # Strip out all navigation, sidebar, footer, category, language elements
            for unwanted in article_body.find_all([
                "div", "span", "ul", "nav"
            ], class_=lambda c: c and any(x in str(c) for x in [
                "navbox","sidebar","reflist","references","catlinks",
                "mw-editsection","toc","hatnote","metadata","infobox",
                "noprint","navigation","footer","langlinks","vector-menu",
                "mw-jump","mw-indicators","printfooter","mw-category",
            ])):
                unwanted.decompose()

            # 1. Wikitables — best source, most structured
            for table in article_body.find_all("table", class_="wikitable"):
                for row in table.find_all("tr"):
                    for cell in row.find_all(["td","th"]):
                        # Linked text in cells = company names
                        for link in cell.find_all("a", href=True):
                            href = link.get("href","")
                            text = link.get_text(strip=True)
                            # Only internal wiki article links, not file/help/category
                            if (text and len(text) > 3 and len(text) < 80
                                    and not text.startswith("[")
                                    and "/wiki/" in href
                                    and not any(x in href for x in [
                                        "File:","Help:","Wikipedia:","Template:",
                                        "Category:","Special:","Talk:","User:",
                                    ])):
                                wiki_names.append(text)
                        # Plain cell text (non-linked)
                        cell_text = cell.get_text(strip=True)
                        if cell_text and len(cell_text) > 3 and len(cell_text) < 60:
                            wiki_names.append(cell_text)

            # 2. Article body lists — but only links to wiki article pages
            for li in article_body.find_all("li"):
                for link in li.find_all("a", href=True):
                    href = link.get("href","")
                    text = link.get_text(strip=True)
                    if (text and len(text) > 3 and len(text) < 80
                            and "/wiki/" in href
                            and not any(x in href for x in [
                                "File:","Help:","Wikipedia:","Template:",
                                "Category:","Special:","Talk:","User:",
                                "#","disambiguation",
                            ])):
                        wiki_names.append(text)

            # Deduplicate
            seen_w = set()
            for name in wiki_names:
                name = name.strip()
                if name and name not in seen_w and len(name) > 3:
                    seen_w.add(name)
                    companies.append(name)

            self.think(f"Wikipedia extracted {len(companies)} entries (article body only)")
            return companies

        # ── General page extraction ──
        # Priority: tables and lists (most structured data)
        structured_text = ""

        # Tables first
        for table in soup.find_all("table")[:3]:
            for row in table.find_all("tr"):
                for cell in row.find_all(["td","th"]):
                    structured_text += cell.get_text(strip=True) + " | "
                structured_text += "\n"

        # Then lists
        for lst in soup.find_all(["ul","ol"])[:5]:
            for li in lst.find_all("li"):
                structured_text += li.get_text(strip=True) + "\n"

        # Fall back to general text
        if len(structured_text) < 200:
            structured_text = soup.get_text(separator=" ", strip=True)[:4000]

        names = self.extract_company_names_from_text(structured_text, context)
        companies.extend(names)
        self.think(f"Extracted {len(names)} company names from page")

        return companies

    def parse_mission(self, mission):
        """
        Break a natural language mission into structured search components.
        e.g. "UK companies completed funding round £100m 2025"
        -> {country: "UK", entity_type: "companies that raised funding", money: "£100m", year: "2025"}
        """
        self.think("🧠 Parsing mission into search components...")

        year_match = re.search(r'\b(20\d{2})\b', mission)
        year = year_match.group(1) if year_match else ""

        money_match = re.search(r'[\$£€]?\d+[\d,]*\s*(?:million|billion|trillion|[MBTmbt])?', mission)
        money = money_match.group(0).strip() if money_match else ""

        country_map = {
            "uk": "UK", "britain": "UK", "british": "UK", "england": "UK",
            "uae": "UAE", "dubai": "UAE", "abu dhabi": "UAE",
            "us ": "US", "usa": "US", "america": "US", "american": "US",
            "europe": "Europe", "european": "Europe",
            "singapore": "Singapore", "hong kong": "Hong Kong",
            "australia": "Australia", "canada": "Canada", "germany": "Germany",
            "ftse": "UK", "lse": "UK",
        }
        country = ""
        for kw, val in country_map.items():
            if kw in mission.lower():
                country = val
                break

        entity_keywords = {
            "funding round": "companies that raised funding",
            "raised": "companies that raised funding",
            "ipo": "companies that IPO'd",
            "listed": "publicly listed companies",
            "family office": "family offices",
            "hedge fund": "hedge funds",
            "private equity": "private equity firms",
            "asset manager": "asset managers",
            "venture capital": "venture capital firms",
            "real estate": "real estate companies",
        }
        entity_type = "companies"
        matched_keyword = ""
        for kw, label in entity_keywords.items():
            if kw in mission.lower():
                entity_type = label
                matched_keyword = kw
                break

        self.think(f"Parsed → country={country} | entity={entity_type} | money={money} | year={year}")
        return {
            "country": country,
            "entity_type": entity_type,
            "money": money,
            "year": year,
            "original": mission,
        }

    def build_registry_queries(self, parsed, country_hint):
        """Build targeted registry queries from parsed mission components"""
        c = parsed["country"] or country_hint
        money = parsed["money"]
        year = parsed["year"]
        entity = parsed["entity_type"]

        size_phrase = f"{money} " if money else ""
        year_phrase = f"{year} " if year else ""
        base = f"{size_phrase}{entity}"

        queries = []
        if "uk" in c.lower():
            queries = [
                f"UK {base} {year_phrase}site:companieshouse.gov.uk",
                f"Companies House UK {base} {year_phrase}",
                f"UK FCA regulated {entity} {size_phrase}{year_phrase}",
                f"site:fca.org.uk {entity} {size_phrase}register",
                f"UK {entity} {size_phrase}{year_phrase}announcement",
            ]
        elif "uae" in c.lower():
            # Check if mission is crypto — if so use crypto-specific UAE registries
            is_crypto = any(w in parsed.get("original","").lower() for w in
                           ["crypto","blockchain","web3","defi","token","digital asset","nft"])
            if is_crypto:
                queries = [
                    f"VARA licensed crypto firms Dubai {year_phrase}",
                    f"DMCC crypto licence holders Dubai {year_phrase}",
                    f"ADGM registered virtual asset {entity} {year_phrase}",
                    f"Dubai {entity} blockchain {size_phrase}{year_phrase}",
                    f"site:crunchbase.com/organization Dubai blockchain {year_phrase}",
                    f"UAE crypto startup {size_phrase}list {year_phrase}",
                ]
            else:
                queries = [
                    f"DFSA licensed {entity} Dubai {size_phrase}{year_phrase}",
                    f"ADGM registered {entity} {size_phrase}{year_phrase}",
                    f"DMCC member {entity} Dubai {year_phrase}",
                    f"Dubai {base} {year_phrase}",
                    f"UAE {entity} {size_phrase}directory {year_phrase}",
                    f"Abu Dhabi {entity} {size_phrase}{year_phrase}",
                ]
        elif "us" in c.lower():
            queries = [
                f"US {base} {year_phrase}site:sec.gov",
                f"SEC registered {entity} {size_phrase}{year_phrase}",
                f"US {base} {year_phrase}",
            ]
        else:
            queries = [
                f"{c} {base} {year_phrase}",
                f"{c} {entity} {size_phrase}list {year_phrase}",
                f"{c} financial regulator registered {entity} {size_phrase}",
            ]

        return [re.sub(r'\s+', ' ', q).strip() for q in queries]

    def search_companies_house(self, country_hint, mission):
        """Search official company registries using smart parsed queries"""
        parsed = self.parse_mission(mission)
        registry_queries = self.build_registry_queries(parsed, country_hint)

        self.think(f"🏛️ Registry search — {len(registry_queries)} targeted queries")
        companies = []
        for query in registry_queries:
            self.think(f"Registry search: {query}")
            results = self.search(query, max_results=5)
            for r in results:
                names = self.extract_company_names_from_text(
                    r.get("content", "") + " " + r.get("title", ""),
                    mission
                )
                companies.extend(names)
            time.sleep(0.5)

        return companies

    def search_document_files(self, mission):
        """Hunt for Excel/CSV/PDF files with company lists"""
        self.think("📄 Hunting document files with company lists...")
        # Use short parsed query for better results
        parsed = self.parse_mission(mission)
        short = f"{parsed.get('country','')} {parsed.get('money','')} {parsed.get('entity','')} {parsed.get('year','')}".strip()
        if len(short) < 5:
            short = mission[:50]
        self.think(f"Document search using: '{short}'")
        doc_queries = [
            f'{short} filetype:xlsx',
            f'{short} filetype:csv list',
            f'{short} filetype:pdf directory',
            f'{short} "download" list excel',
            f'site:github.com {short} companies csv',
            f'site:github.com {short} fund list json',
            f'site:data.gov.uk {short}',
        ]

        companies = []
        for query in doc_queries:
            self.think(f"Document search: {query}")
            results = self.search(query, max_results=5)
            for r in results:
                url = r.get("url","")
                content = r.get("content","") + " " + r.get("title","")
                names = self.extract_company_names_from_text(content, mission)
                companies.extend(names)
                # Actually READ document files!
                if any(url.lower().endswith(ext) for ext in ['.xlsx','.xls']):
                    self.think(f"📊 Reading Excel: {url}")
                    companies.extend(self.read_excel_url(url))
                elif url.lower().endswith('.csv'):
                    self.think(f"📋 Reading CSV: {url}")
                    companies.extend(self.read_csv_url(url))
                elif url.lower().endswith('.pdf'):
                    # Check file size before downloading — skip huge PDFs
                    try:
                        head = requests.head(url, timeout=4, headers=HEADERS, allow_redirects=True)
                        size = int(head.headers.get("content-length", 0))
                        if size > 5_000_000:
                            self.think(f"⏭️ Skipping {size//1_000_000}MB PDF — too large")
                            continue
                    except: pass
                    self.think(f"📄 Reading PDF: {url}")
                    companies.extend(self.read_pdf_url(url))
            time.sleep(0.3)

        return companies

    def search_wikipedia_lists(self, mission):
        """Wikipedia has amazing lists of companies — hits high-value pages directly"""
        self.think("📚 Searching Wikipedia lists...")
        companies = []
        seen_urls = set()

        # Detect mission type for direct page targeting
        mission_lower = mission.lower()
        direct_pages = []

        if any(w in mission_lower for w in ["private equity","pe firm","buyout"]):
            direct_pages += [
                "https://en.wikipedia.org/wiki/List_of_private_equity_firms",
                "https://en.wikipedia.org/wiki/Category:Private_equity_firms_of_the_United_Kingdom",
            ]
        if any(w in mission_lower for w in ["hedge fund","hedge"]):
            direct_pages += [
                "https://en.wikipedia.org/wiki/List_of_hedge_funds",
                "https://en.wikipedia.org/wiki/Category:Hedge_funds",
            ]
        if any(w in mission_lower for w in ["family office","wealth","family"]):
            direct_pages += [
                "https://en.wikipedia.org/wiki/Family_office",
                "https://en.wikipedia.org/wiki/Category:Investment_management_companies_of_the_United_Kingdom",
            ]
        if any(w in mission_lower for w in ["ftse","listed","stock exchange","london stock"]):
            direct_pages += [
                "https://en.wikipedia.org/wiki/FTSE_250_Index",
                "https://en.wikipedia.org/wiki/FTSE_100_Index",
                "https://en.wikipedia.org/wiki/Category:Companies_in_the_FTSE_250_Index",
            ]
        if any(w in mission_lower for w in ["asset management","asset manager","investment management"]):
            direct_pages += [
                "https://en.wikipedia.org/wiki/Category:Investment_management_companies_of_the_United_Kingdom",
                "https://en.wikipedia.org/wiki/List_of_asset_management_firms",
            ]
        if any(w in mission_lower for w in ["sovereign","wealth fund","pension"]):
            direct_pages += [
                "https://en.wikipedia.org/wiki/List_of_largest_pension_funds",
                "https://en.wikipedia.org/wiki/Sovereign_wealth_fund",
            ]

        if any(w in mission_lower for w in ["dubai","uae","emirates","abu dhabi"]):
            direct_pages += [
                "https://en.wikipedia.org/wiki/List_of_companies_of_the_United_Arab_Emirates",
                "https://en.wikipedia.org/wiki/Dubai_International_Financial_Centre",
                "https://en.wikipedia.org/wiki/Abu_Dhabi_Global_Market",
            ]
        if any(w in mission_lower for w in ["crypto","blockchain","web3","defi","bitcoin"]):
            direct_pages += [
                "https://en.wikipedia.org/wiki/List_of_cryptocurrency_exchanges",
                "https://en.wikipedia.org/wiki/Decentralized_finance",
            ]

        # Always add general UK investment page
        if any(w in mission_lower for w in ["uk","britain","london","british"]):
            direct_pages.append("https://en.wikipedia.org/wiki/Category:Investment_management_companies_of_the_United_Kingdom")

        # Hit direct pages first
        for url in direct_pages:
            if url in seen_urls: continue
            seen_urls.add(url)
            results = self.hunt_from_url(url, mission)
            companies.extend(results)
            self.think(f"Direct page: {len(results)} companies from {url.split('/')[-1]}")

        # Then search for more
        wiki_queries = [
            f"site:en.wikipedia.org list of {mission}",
            f"site:en.wikipedia.org {mission} firms",
        ]
        for query in wiki_queries:
            results = self.search(query, max_results=3)
            for r in results:
                url = r.get("url","")
                if "wikipedia.org" in url and url not in seen_urls:
                    seen_urls.add(url)
                    companies.extend(self.hunt_from_url(url, mission))
            time.sleep(0.3)

        return companies

    def hunt_community_sources(self, mission):
        """
        Hunt Discord servers, Telegram groups, Reddit communities, and other
        social/community platforms for company and project names.

        Returns list of dicts: {name, source_type, source_url, source_name}
        so the community link is preserved all the way into the CRM.
        """
        self.think("🌐 Community hunt: Discord / Telegram / Reddit / Forums...")
        mission_type = self.classify_mission(mission)
        # Results: list of dicts with company name + source metadata
        results_out = []
        # Track community sources found (url → metadata) for storage
        self._community_sources_found = []

        try:
            parsed = self.parse_mission(mission)
            country = parsed.get("country","")
            entity  = parsed.get("entity_type","companies")
        except:
            country = ""
            entity  = "companies"

        def add_result(name, source_type, source_url, source_name):
            """Add a company name with its community source metadata"""
            if not name or not isinstance(name, str) or len(name.strip()) < 4:
                return
            results_out.append({
                "name":        name.strip(),
                "source_type": source_type,
                "source_url":  source_url,
                "source_name": source_name,
            })

        def extract_and_add(text, url, source_type, source_name):
            """Extract company names from text and tag each with source metadata"""
            names = self.extract_company_names_from_text(text, mission)
            for n in names:
                add_result(n, source_type, url, source_name)

        def register_community(url, source_type, name, description=""):
            """Register a community URL for storage in company_sources later"""
            if url and url not in [c["url"] for c in self._community_sources_found]:
                self._community_sources_found.append({
                    "source_type": source_type,
                    "url":         url,
                    "name":        name,
                    "notes":       description,
                })

        # ── Discord ──────────────────────────────────────────────
        self.think("💬 Searching Discord community directories...")

        # Extract the topic from the mission — strip community trigger words
        topic_words = re.sub(
            r"\b(discord|telegram|reddit|channel|server|group|community|communities|find|search|hunt|list|groups|servers|channels)\b",
            "", mission, flags=re.IGNORECASE
        ).strip()
        topic = topic_words if len(topic_words) > 2 else entity

        # Always search the main Discord server directories directly
        discord_direct_pages = [
            f"https://disboard.org/servers/tag/{quote_plus(topic)}",
            f"https://discordservers.com/servers/{quote_plus(topic)}",
        ]
        for page_url in discord_direct_pages:
            self.think(f"💬 Discord directory: {page_url}")
            page_companies = self.hunt_from_url(page_url, mission)
            for item in (page_companies or []):
                n = item.get("name","") if isinstance(item, dict) else item
                add_result(n, "discord", page_url, f"Discord directory: {topic}")
            register_community(page_url, "discord", f"Discord servers — {topic}")
            time.sleep(0.5)

        if mission_type == "crypto" or "crypto" in mission.lower():
            discord_queries = [
                f"site:disboard.org crypto {topic} discord server",
                f"site:discordservers.com crypto {topic}",
                f"site:discord.me crypto {topic} server",
                f"discord.gg crypto {topic} server invite 2024",
                f"best crypto discord servers {topic} {country} 2024",
            ]
        elif mission_type in ("startup","pe_vc"):
            discord_queries = [
                f"site:disboard.org startup founders {topic} discord server",
                f"site:discordservers.com startups {topic}",
                f"discord {country} startup founders {topic} server 2024",
            ]
        else:
            discord_queries = [
                f"site:disboard.org {topic} {country} discord server",
                f"site:discordservers.com {topic} {country}",
                f"site:discord.me {topic} community server",
            ]

        for query in discord_queries:
            self.think(f"Discord: {query}")
            search_results = self.search(query, max_results=5)
            for r in search_results:
                url   = r.get("url","")
                title = r.get("title","")
                text  = r.get("content","") + " " + title
                # Register any discord.gg or discord.com links as community sources
                discord_links = re.findall(r'https?://(?:discord\.gg|discord\.com/invite)/[\w-]+', text)
                for link in discord_links:
                    register_community(link, "discord", title or "Discord Server")
                if "discord.gg" in url or "discord.com/invite" in url:
                    register_community(url, "discord", title or "Discord Server")
                extract_and_add(text, url, "discord", title or "Discord community")
                # Deep dive disboard pages
                if "disboard.org" in url:
                    register_community(url, "discord", title or "Disboard listing")
                    page_companies = self.hunt_from_url(url, mission)
                    for item in (page_companies or []):
                        n = item.get("name","") if isinstance(item, dict) else item
                        add_result(n, "discord", url, title or "Disboard")
            time.sleep(0.3)

        # ── Telegram ─────────────────────────────────────────────
        self.think("📱 Searching Telegram groups...")

        # Always search Telegram directories directly
        telegram_direct_pages = [
            f"https://telemetr.io/en/channels?search={quote_plus(topic)}",
            f"https://tgstat.com/search?q={quote_plus(topic)}",
        ]
        for page_url in telegram_direct_pages:
            self.think(f"📱 Telegram directory: {page_url}")
            page_companies = self.hunt_from_url(page_url, mission)
            for item in (page_companies or []):
                n = item.get("name","") if isinstance(item, dict) else item
                add_result(n, "telegram", page_url, f"Telegram directory: {topic}")
            register_community(page_url, "telegram", f"Telegram channels — {topic}")
            time.sleep(0.5)

        if mission_type == "crypto" or "crypto" in mission.lower():
            telegram_queries = [
                f"site:t.me crypto {topic} channel",
                f"site:telemetr.io crypto {topic} {country} channel",
                f"best crypto telegram channels {topic} {country} 2024",
                f"telegram crypto {topic} group invite link {country}",
                f"t.me crypto {topic} blockchain community {country}",
            ]
        elif mission_type == "family_office":
            telegram_queries = [
                f"site:t.me family office wealth {country} channel",
                f"telegram investment wealth management {country} group",
            ]
        else:
            telegram_queries = [
                f"site:t.me {topic} {country} channel",
                f"site:telemetr.io {topic} {country}",
                f"telegram {topic} {country} group community invite",
            ]

        for query in telegram_queries:
            self.think(f"Telegram: {query}")
            search_results = self.search(query, max_results=5)
            for r in search_results:
                url   = r.get("url","")
                title = r.get("title","")
                text  = r.get("content","") + " " + title
                # Register t.me links
                tg_links = re.findall(r'https?://t\.me/[\w-]+', text)
                for link in tg_links:
                    register_community(link, "telegram", title or "Telegram Channel")
                if "t.me/" in url:
                    register_community(url, "telegram", title or "Telegram Channel")
                extract_and_add(text, url, "telegram", title or "Telegram community")
                # Deep dive Telegram directory pages
                if any(s in url for s in ["telemetr.io","tgstat.com","telegramchannels.me"]):
                    register_community(url, "telegram", title or "Telegram directory")
                    page_companies = self.hunt_from_url(url, mission)
                    for item in (page_companies or []):
                        n = item.get("name","") if isinstance(item, dict) else item
                        add_result(n, "telegram", url, title or "Telegram directory")
            time.sleep(0.3)

        # ── Reddit ───────────────────────────────────────────────
        self.think("🔴 Searching Reddit communities...")

        if mission_type == "crypto":
            reddit_queries = [
                f"site:reddit.com {country} crypto startup companies list",
                f"reddit {country} blockchain web3 companies subreddit",
                f"site:reddit.com/r/CryptoCurrency {country} projects list",
                f"site:reddit.com {country} crypto startups to watch 2024",
            ]
        elif mission_type in ("startup","pe_vc"):
            reddit_queries = [
                f"site:reddit.com {country} startups list companies",
                f"site:reddit.com/r/startups {country} companies",
            ]
        elif mission_type == "family_office":
            reddit_queries = [
                f"site:reddit.com family office {country} firms list",
            ]
        else:
            reddit_queries = [
                f"site:reddit.com {country} {entity} companies list",
                f"reddit {country} {entity} firms recommendations",
            ]

        for query in reddit_queries:
            self.think(f"Reddit: {query}")
            search_results = self.search(query, max_results=5)
            for r in search_results:
                url   = r.get("url","")
                title = r.get("title","")
                text  = r.get("content","") + " " + title
                if "reddit.com/r/" in url:
                    register_community(url, "reddit", title or "Reddit thread")
                extract_and_add(text, url, "reddit", title or "Reddit community")
            time.sleep(0.3)

        # ── Mission-specific community platforms ─────────────────
        self.think("🏛️ Searching mission-specific community platforms...")

        if mission_type == "crypto":
            crypto_community_queries = [
                (f"{country} crypto ecosystem map projects 2024",          "web",      "Ecosystem map"),
                (f"site:defillama.com {country} protocols",                "defillama","DeFiLlama"),
                (f"site:coingecko.com {country} {entity} list",            "coingecko","CoinGecko"),
                (f"site:coinmarketcap.com {country} projects",             "coinmktcap","CoinMarketCap"),
                (f"Token2049 Dubai sponsors exhibitors 2024",              "event",    "Token2049 Dubai"),
                (f"Future Blockchain Summit Dubai 2024 companies",         "event",    "Future Blockchain Summit"),
                (f"ETH Dubai 2024 sponsors projects",                      "event",    "ETH Dubai"),
                (f"Gitcoin grants {country} projects",                     "gitcoin",  "Gitcoin Grants"),
                (f"Binance Labs portfolio {country} companies",            "vc_port",  "Binance Labs portfolio"),
                (f"a16z crypto portfolio {country} companies",             "vc_port",  "a16z crypto portfolio"),
            ]
            for query, src_type, src_name in crypto_community_queries:
                self.think(f"Crypto community: {query}")
                search_results = self.search(query, max_results=5)
                for r in search_results:
                    url  = r.get("url","")
                    text = r.get("content","") + " " + r.get("title","")
                    extract_and_add(text, url, src_type, src_name)
                    if any(s in url for s in ["defillama","coingecko","coinmarketcap"]):
                        register_community(url, src_type, src_name)
                        page_companies = self.hunt_from_url(url, mission)
                        for item in (page_companies or []):
                            n = item.get("name","") if isinstance(item, dict) else item
                            add_result(n, src_type, url, src_name)
                time.sleep(0.3)

        elif mission_type in ("startup","pe_vc"):
            startup_community_queries = [
                (f"ProductHunt {country} top companies 2024",      "producthunt", "ProductHunt"),
                (f"AngelList {country} startups list",             "angellist",   "AngelList"),
                (f"Y Combinator {country} companies batch",        "yc",          "Y Combinator"),
                (f"Hacker News {country} startup launches",        "hackernews",  "Hacker News"),
            ]
            for query, src_type, src_name in startup_community_queries:
                self.think(f"Startup community: {query}")
                search_results = self.search(query, max_results=4)
                for r in search_results:
                    url  = r.get("url","")
                    text = r.get("content","") + " " + r.get("title","")
                    extract_and_add(text, url, src_type, src_name)
                time.sleep(0.3)

        self.think(f"🌐 Community hunt complete: {len(results_out)} company names found across {len(self._community_sources_found)} community sources")
        return results_out

    def find_companies(self, mission, target_count=50):
        """
        Main deep hunt — finds companies one by one from many sources
        Returns list of unique company names
        NOTE: target_count is a MINIMUM — JAMIE hunts everything she finds
        """
        # Internal limit is much larger — we never stop early
        # If we find 250 when asked for 20, we return all 250
        internal_limit = max(target_count * 20, 1000)

        self.think(f"\n🎯 DEEP HUNT MODE")
        self.think(f"Mission: {mission}")
        self.think(f"Target: {target_count} minimum — JAMIE hunts everything she finds")

        # ── Community-first routing ───────────────────────────
        # If the mission is about finding Discord/Telegram/Reddit communities,
        # skip wiki/registry/document sources entirely — go straight to the
        # community hunter which targets disboard.org, t.me, reddit, etc.
        if self.classify_mission(mission) == "community":
            self.think(f"Strategy: Community hunt mode — searching Discord / Telegram / Reddit\n")
            community_results = self.hunt_community_sources(mission)
            _seen = set()
            all_communities = []
            for item in community_results:
                name = item.get("name","") if isinstance(item, dict) else str(item)
                if name and name.strip() not in _seen and len(name.strip()) >= 4:
                    _seen.add(name.strip())
                    all_communities.append(name.strip())
            self._last_community_sources = getattr(self, "_community_sources_found", [])
            self.think(f"\n✅ Community hunt complete: {len(all_communities)} names found from {len(self._last_community_sources)} community sources")
            return all_communities

        self.think(f"Strategy: Multi-layer deep search\n")

        all_companies = []
        seen = set()

        def add_company(name, location=None, aum=None, company_type=None, source_url=None):
            """Add company with whatever intel we already have from the source page"""
            if not name or not isinstance(name, str): return False
            name = name.strip()
            if not name or name in seen or len(name) < 4: return False
            if name.startswith("[DOC:"): return False
            # Skip generic words
            generic = ["company","fund","office","investment","group","limited","the","and","of",
                       "article","talk","read","edit","contents","help","donate","category",
                       "main page","log in","create account","privacy policy","contact"]
            if name.lower() in generic: return False
            # Skip Wikipedia nav items
            if name.lower().startswith(("see also","references","external","notes","further")):
                return False
            # Skip pure numbers, years, cities alone
            if re.match(r'^[\d\s\.,\-]+$', name): return False
            seen.add(name)
            record = {
                "name": name,
                "location": location,
                "aum": aum,
                "company_type": company_type,
                "source_url": source_url,
            }
            all_companies.append(record)
            self.think(f"✅ Found [{len(all_companies)}]: {name}" + (f" | {location}" if location else "") + (f" | {aum}" if aum else ""))
            return True

        # Detect country from mission
        country_hint = ""
        for c in ["UK","UAE","US","USA","Europe","Asia","Middle East","London","Dubai","New York","FTSE","LSE","Britain","British"]:
            if c.lower() in mission.lower():
                if c in ["FTSE","LSE","Britain","British","London"]:
                    country_hint = "UK"
                else:
                    country_hint = c
                break

        # ── Source 1: Official registries ──────────────────
        self.think("\n📋 Source 1: Official registries...")
        if country_hint:
            registry_companies = self.search_companies_house(country_hint, mission)
            for name in registry_companies:
                add_company(name, source_url="registry")
                if len(all_companies) >= internal_limit: break

        # ── Source 2: Wikipedia lists ───────────────────────
        if len(all_companies) < internal_limit:
            self.think("\n📚 Source 2: Wikipedia lists...")
            wiki_companies = self.search_wikipedia_lists(mission)
            for name in wiki_companies:
                if isinstance(name, dict):
                    add_company(name.get("name",""), location=name.get("location"), aum=name.get("aum"), source_url="wikipedia")
                else:
                    add_company(name, source_url="wikipedia")
                if len(all_companies) >= internal_limit: break

        # ── Source 3: Document files ────────────────────────
        # Skip if Wikipedia already found plenty of quality targets
        if len(all_companies) >= 200:
            self.think(f"✅ Already have {len(all_companies)} targets from Wikipedia — skipping document hunt")
        elif len(all_companies) < internal_limit:
            self.think("\n📄 Source 3: Document hunting (Excel/CSV/PDF)...")
            doc_companies = self.search_document_files(mission)
            for name in doc_companies:
                if isinstance(name, dict):
                    add_company(name.get("name",""), location=name.get("location"), aum=name.get("aum"), source_url="document")
                else:
                    add_company(name, source_url="document")
                if len(all_companies) >= internal_limit: break
        # end Source 3

        # ── Source 4: Brain-generated deep queries ──────────
        if len(all_companies) < internal_limit:
            self.think("\n🧠 Source 4: Brain-generated deep queries...")
            deep_queries = self.generate_deep_queries(mission)

            for i, query in enumerate(deep_queries):
                if len(all_companies) >= internal_limit: break
                self.think(f"Query [{i+1}/{len(deep_queries)}]: {query}")

                results = self.search(query, max_results=10)
                for r in results:
                    content = r.get("content","") + " " + r.get("title","")
                    url = r.get("url","")

                    # Extract names from search snippet
                    names = self.extract_company_names_from_text(content, mission)
                    for name in names:
                        add_company(name, source_url=url)

                    # Deep dive into promising pages
                    if len(all_companies) < internal_limit and any(kw in url.lower() for kw in ["list","directory","top","ranking","register","members"]):
                        page_companies = self.hunt_from_url(url, mission)
                        for item in page_companies:
                            if isinstance(item, dict):
                                add_company(item.get("name",""), location=item.get("location"), aum=item.get("aum"), source_url=url)
                            else:
                                add_company(item, source_url=url)

                    if len(all_companies) >= internal_limit: break
                time.sleep(0.5)

        # ── Source 5: Industry-specific searches ────────────
        if len(all_companies) < internal_limit:
            self.think("\n🏭 Source 5: Industry-specific searches...")
            mission_type = self.classify_mission(mission)
            industry_queries = ask_jamie(f"""Generate 10 very specific search queries to find {mission_type} companies matching: "{mission}"
Focus on: trade associations, industry bodies, conference lists, award winners, regulatory filings.
Each query must be 4-8 words and directly target companies that match the mission.
Return ONLY a JSON array.""", max_tokens=300, temperature=0.7)

            try:
                match = re.search(r'\[.*?\]', industry_queries, re.DOTALL)
                if match:
                    queries = json.loads(match.group())
                    for query in queries[:10]:
                        if len(all_companies) >= internal_limit: break
                        if isinstance(query, str):
                            self.think(f"Industry: {query}")
                            for r in self.search(query, max_results=8):
                                text = r.get("content","") + " " + r.get("title","")
                                for name in self.extract_company_names_from_text(text, mission):
                                    add_company(name)
                            time.sleep(0.5)
            except: pass

        # ── Source 6: Community sources (Discord/Telegram/Reddit) ──
        if len(all_companies) < internal_limit:
            self.think("\n🌐 Source 6: Community sources (Discord / Telegram / Reddit)...")
            community_results = self.hunt_community_sources(mission)
            for item in community_results:
                if isinstance(item, dict):
                    name = item.get("name","")
                    src  = item.get("source_url","community")
                    add_company(name, source_url=src)
                else:
                    add_company(item, source_url="community")
                if len(all_companies) >= internal_limit: break
            # Attach community sources list to searcher for push_to_crm to use
            self._last_community_sources = getattr(self, "_community_sources_found", [])

        self.think(f"\n✅ Deep hunt complete: Found {len(all_companies)} unique companies")
        self.think(f"\n✅ Found {len(all_companies)} companies — hunting ALL of them!")
        return all_companies  # JAMIE never leaves a lead behind


# ── EXECUTIVE HUNTER ──────────────────────────────────────────
# (kept from v4 — working well for single companies)

class JAMIEHunter:
    def __init__(self):
        self.log = []
        self.searcher = DeepSearcher()

    def think(self, msg):
        timestamp = datetime.now().strftime("%H:%M:%S")
        entry = f"[{timestamp}] {msg}"
        self.log.append(entry)
        print(f"  🕵️  JAMIE: {msg}")

    def search(self, query, max_results=5):
        return self.searcher.search(query, max_results)

    def fetch_page(self, url):
        _, soup = self.searcher.fetch_page(url)
        return soup

    def extract_emails(self, text):
        pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        return [e for e in re.findall(pattern, text)
                if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', e)]

    def mx_lookup(self, domain):
        try:
            import dns.resolver
            answers = dns.resolver.resolve(domain, "MX")
            providers = [str(r.exchange).lower() for r in answers]
            if any("google" in p for p in providers): return "Google Workspace"
            if any("microsoft" in p or "outlook" in p for p in providers): return "Microsoft 365"
            return "Custom"
        except:
            return None

    def parse_value(self, text):
        if not text: return None
        text = str(text).strip().replace(",","").replace("$","").replace("£","").replace("€","")
        for word, mult in [("trillion",1e12),("billion",1e9),("million",1e6),("thousand",1e3)]:
            m = re.search(rf'([\d]+\.?[\d]*)\s*{word}', text.lower())
            if m: return float(m.group(1)) * mult
        for letter, mult in [("t",1e12),("b",1e9),("m",1e6),("k",1e3)]:
            m = re.search(rf'([\d]+\.?[\d]*)\s*{letter}\b', text.lower())
            if m: return float(m.group(1)) * mult
        m = re.search(r'[\d]+\.?[\d]*', text)
        return float(m.group()) if m else None

    def get_domain(self, company_name):
        """Smart domain detection — tries multiple strategies"""
        self.think(f"Finding domain: {company_name}")

        # Skip obvious non-company entries
        skip_words = ["financial services authority","international financial centre",
                      "stock exchange","companies house","ernst & young","deloitte",
                      "pwc","kpmg","regulatory authority","government"]
        if any(s in company_name.lower() for s in skip_words):
            self.think(f"⚠️ Skipping regulatory/gov entity: {company_name}")
            return None

        # Strategy 1: Direct search for official site
        for query in [
            f'"{company_name}" official website',
            f'{company_name} site',
            f'{company_name} homepage',
        ]:
            for r in self.search(query, max_results=5):
                url = r.get("url","")
                if not url: continue
                domain = urlparse(url).netloc.replace("www.","")
                # Skip social media, directories, news sites
                skip_domains = ["linkedin.com","facebook.com","twitter.com","bloomberg.com",
                               "reuters.com","crunchbase.com","wikipedia.org","glassdoor.com",
                               "indeed.com","companies","companieshouse","gov.","regulatory"]
                if any(s in domain for s in skip_domains): continue
                if domain and "." in domain and len(domain) > 4:
                    self.think(f"Domain: {domain}")
                    return domain

        # Strategy 2: Brain guesses the domain
        guess = ask_jamie(f"""What is the official website domain of "{company_name}"?
Return ONLY the domain like: company.com or company.ae or company.co.uk
No http, no www, just the domain.
Domain:""", max_tokens=20, temperature=0.1).strip().lower()
        guess = re.sub(r'[^a-z0-9\.\-]', '', guess)
        if guess and "." in guess and len(guess) > 4:
            self.think(f"Brain guessed domain: {guess}")
            return guess

        # Strategy 3: Fallback pattern
        clean = re.sub(r'[^a-z0-9]','', company_name.lower().split("(")[0])
        return f"{clean}.com" if clean else None

    def get_role_aliases(self, role):
        """Map role names to search terms and title variations"""
        aliases = {
            "CEO":  ["CEO","Chief Executive","Managing Director","President","Head of"],
            "CFO":  ["CFO","Chief Financial Officer","Finance Director","Head of Finance"],
            "COO":  ["COO","Chief Operating Officer","Operations Director","Head of Operations"],
            "CMO":  ["CMO","Chief Marketing Officer","Marketing Director","Head of Marketing","VP Marketing"],
            "CTO":  ["CTO","Chief Technology Officer","Technology Director","Head of Technology"],
            "CSR":  ["CSR Director","Head of CSR","Corporate Responsibility","Sustainability Director",
                     "Head of Sustainability","ESG Director","Social Impact"],
            "Director": ["Director","Board Member","Non-Executive Director","Executive Director"],
            "Partner":  ["Partner","Managing Partner","Senior Partner","Founding Partner"],
        }
        # Find best match
        role_upper = role.upper()
        for key, vals in aliases.items():
            if role_upper == key or role_upper in [v.upper() for v in vals]:
                return vals
        # Return role as-is if no match
        return [role]

    def hunt_executive(self, company_name, domain, role="CEO"):
        """
        Hunt an executive by role — flexible, handles any role title.
        Works for CEO, CFO, COO, CMO, CSR Director, Marketing Director etc.
        """
        self.think(f"Hunting {role} at {company_name}...")
        result = {"role":role,"name":None,"email":None,"linkedin":None,
                  "confidence":0,"method":None,"notes":[]}

        # Skip regulatory/gov entities
        skip_words = ["financial services authority","stock exchange","companies house",
                      "regulatory authority","government of","ministry of"]
        if any(s in company_name.lower() for s in skip_words):
            self.think(f"⚠️ Skipping regulatory entity: {company_name}")
            result["name"] = f"[{role} — Not Applicable]"
            return result

        aliases = self.get_role_aliases(role)
        primary_alias = aliases[0]

        # ── Layer 1: Multiple search queries ──────────────────
        search_text = ""
        search_queries = [
            f"{company_name} {primary_alias}",
            f"{company_name} {role} name",
            f'"{company_name}" {primary_alias} 2024 2025',
            f'{company_name} leadership team {role}',
        ]
        for query in search_queries:
            if result["name"]: break
            for r in self.search(query, max_results=5):
                content = r.get("title","") + " " + r.get("content","")
                search_text += content + " "
                # Try each alias in patterns
                for alias in aliases:
                    for pattern in [
                        rf'([A-Z][a-z]{{2,}} [A-Z][a-z]{{2,}})(?:\s+is|\s+serves|\s+appointed)?\s+(?:as\s+)?(?:the\s+)?{re.escape(alias)}',
                        rf'{re.escape(alias)}[,\s]+([A-Z][a-z]{{2,}} [A-Z][a-z]{{2,}})',
                    ]:
                        try:
                            m = re.search(pattern, content, re.IGNORECASE)
                            if m:
                                result["name"] = m.group(1)
                                result["confidence"] = 45
                                result["method"] = f"Search ({alias})"
                                break
                        except: pass
                    if result["name"]: break
                if result["name"]: break

        # ── Layer 2: Brain reads search results ───────────────
        if not result["name"] and search_text:
            aliases_str = " / ".join(aliases[:3])
            name = ask_jamie(f"""Find the person with role "{role}" (also known as: {aliases_str}) at the company "{company_name}".
Look carefully through this text.
Return ONLY their full name, or NOT_FOUND if genuinely not present.
Text: {search_text[:3000]}
Name:""", max_tokens=40, temperature=0.2).strip()
            # Clean up brain response
            name = name.split("\n")[0].strip().rstrip(".")
            if name and "NOT_FOUND" not in name and len(name.split()) >= 2:
                # Validate looks like a name
                if re.match(r'^[A-Z]', name) and not any(w in name.lower() for w in ["the ","is ","are ","was "]):
                    result["name"] = name
                    result["confidence"] = 52
                    result["method"] = "Brain — search"

        # ── Layer 3: Website scraping ─────────────────────────
        if domain and not result["name"]:
            for path in ["/about","/about-us","/team","/leadership","/management",
                         "/people","/our-team","/executives","/directors","/partners"]:
                soup = self.fetch_page(f"https://{domain}{path}")
                if not soup: continue
                text = soup.get_text()
                # Check page has relevant content
                if not any(a.lower() in text.lower() for a in aliases): continue
                name = ask_jamie(f"""Find the {role} (or: {", ".join(aliases[:3])}) of {company_name}.
Return ONLY their full name or NOT_FOUND.
Text: {text[:3000]}
Name:""", max_tokens=40, temperature=0.2).strip()
                name = name.split("\n")[0].strip().rstrip(".")
                if name and "NOT_FOUND" not in name and len(name.split()) >= 2:
                    if re.match(r'^[A-Z]', name) and not any(w in name.lower() for w in ["the ","is ","was "]):
                        result["name"] = name
                        result["confidence"] = 62
                        result["method"] = "Website+Brain"
                        break

        # ── Layer 4: LinkedIn — flexible role search ──────────
        if not result["name"] or not result["linkedin"]:
            for alias in aliases[:2]:
                for r in self.search(f'site:linkedin.com "{company_name}" "{alias}"', max_results=3):
                    url = r.get("url","")
                    if "linkedin.com/in/" in url:
                        result["linkedin"] = url
                        # Extract name from LinkedIn title
                        title = r.get("title","")
                        # LinkedIn titles: "John Smith - CEO at Company | LinkedIn"
                        nm = re.match(r'^([A-Z][a-zA-Z\'-]+ [A-Z][a-zA-Z\'-]+)', title)
                        if nm and not result["name"]:
                            result["name"] = nm.group(1)
                            result["confidence"] = 68
                            result["method"] = f"LinkedIn ({alias})"
                        break
                if result["linkedin"]: break

        # ── Layer 5: Appointment news search ─────────────────
        if not result["name"]:
            for alias in aliases[:2]:
                for r in self.search(f'"{company_name}" "{alias}" appointed named 2023 2024 2025', max_results=3):
                    text = r.get("content","") + " " + r.get("title","")
                    name = ask_jamie(f"""Find the full name of the person who is {alias} at {company_name}.
Look for appointment news, press releases, leadership announcements.
Return ONLY their full name (First Last), or NOT_FOUND.
Text: {text[:1500]}
Name:""", max_tokens=40, temperature=0.2).strip()
                    name = name.split("\n")[0].strip().rstrip(".")
                    if name and "NOT_FOUND" not in name and len(name.split()) >= 2:
                        if re.match(r'^[A-Z]', name) and not any(w in name.lower() for w in ["the ","is ","was "]):
                            result["name"] = name
                            result["confidence"] = 48
                            result["method"] = f"Appointment news ({alias})"
                            break
                if result["name"]: break

        # ── Layer 6: Crunchbase people search ─────────────────
        if not result["name"]:
            cb_query = f'site:crunchbase.com/person "{company_name}" {primary_alias}'
            for r in self.search(cb_query, max_results=3):
                url = r.get("url","")
                title = r.get("title","")
                if "crunchbase.com/person" in url:
                    # Crunchbase titles: "John Smith - CEO at Company | Crunchbase"
                    nm = re.match(r'^([A-Z][a-zA-Z\'-]+(?: [A-Z][a-zA-Z\'-]+)+)', title)
                    if nm:
                        result["name"] = nm.group(1)
                        result["confidence"] = 58
                        result["method"] = "Crunchbase"
                        break

        # ── Layer 7: Companies House officers (UK only) ────────
        if not result["name"] and domain and any(ext in domain for ext in [".co.uk",".uk",".ltd"]):
            ch_query = f'site:find-and-update.company-information.service.gov.uk "{company_name}" director officer'
            for r in self.search(ch_query, max_results=3):
                text = r.get("content","") + " " + r.get("title","")
                name = ask_jamie(f"""Find the {primary_alias} or director name of {company_name} from this Companies House data.
Return ONLY their full name or NOT_FOUND.
Text: {text[:1500]}
Name:""", max_tokens=40, temperature=0.2).strip()
                name = name.split("\n")[0].strip().rstrip(".")
                if name and "NOT_FOUND" not in name and len(name.split()) >= 2:
                    if re.match(r'^[A-Z]', name):
                        result["name"] = name
                        result["confidence"] = 60
                        result["method"] = "Companies House"
                        break

        # ── Layer 6: Email generation ─────────────────────────
        if result["name"] and domain:
            parts = result["name"].split()
            if len(parts) >= 2:
                f, l = parts[0].lower(), parts[-1].lower()
                candidates = [
                    f"{f}.{l}@{domain}", f"{f[0]}{l}@{domain}",
                    f"{f}@{domain}", f"{f}{l}@{domain}",
                    f"{l}.{f}@{domain}", f"{f}.{l[0]}@{domain}",
                ]
                mx = self.mx_lookup(domain)
                if mx: result["notes"].append(f"Provider: {mx}")
                result["email"] = candidates[0]
                result["confidence"] = min(result["confidence"]+15, 85)
                self.think(f"Email: {result['email']} ({result['confidence']}%)")

        if not result["name"] or "Manual Review" in (result["name"] or ""):
            result["name"] = f"[{role} — Manual Review]"
            result["confidence"] = 0
            self.think(f"⚠️ {role} not found after all layers")
        else:
            self.think(f"✅ {role}: {result['name']} | {result.get('email','no email')} | {result['confidence']}%")

        return result

        return result

    def hunt_company_profile(self, company_name, domain, pre_intel=None):
        """Build company profile — uses pre_intel if available to skip redundant searches"""
        pre = pre_intel or {}
        profile = {
            "company": company_name, "domain": domain,
            "location": pre.get("location"),
            "founded": None, "employees": None,
            "company_type": pre.get("company_type", "Unknown"),
            "valuation": None,
            "valuation_raw": pre.get("aum"),
            "funding_status": None, "liquidity_flag": "UNKNOWN",
            "qualified": False, "qualification_notes": []
        }
        # If we already have key intel, skip the search phase
        already_have = bool(pre.get("location") or pre.get("aum") or pre.get("company_type"))
        if already_have:
            self.think(f"ℹ️ Using pre-loaded intel for {company_name} — skipping profile search")

        # Only search if we don't already have the intel
        all_text = ""
        if not already_have:
            for r in self.search(f"{company_name} AUM valuation headquarters founded employees")[:5]:
                all_text += r.get("content","") + " " + r.get("title","") + " "
        else:
            all_text = f"{company_name} {pre.get('location','')} {pre.get('aum','')}"

        loc = re.search(r'headquartered? (?:in )?([A-Z][a-zA-Z\s]+,\s*[A-Z][a-zA-Z\s]+)', all_text)
        if loc: profile["location"] = loc.group(1).strip()

        for pattern in [
            r'([\d]+\.?[\d]*\s*(?:trillion|billion|million|[TBM]))\s*(?:in\s+)?(?:AUM|assets)',
            r'(?:AUM|assets under management)[^\d]*([\d]+\.?[\d]*\s*(?:trillion|billion|million|[TBM]))',
            r'market cap[^\d]*([\d]+\.?[\d]*\s*(?:trillion|billion|million|[TBM]))',
            r'\$([\d]+\.?[\d]*\s*(?:trillion|billion|million|[TBM]))',
        ]:
            m = re.search(pattern, all_text, re.IGNORECASE)
            if m:
                raw = m.group(1).strip()
                val = self.parse_value(raw)
                if val and val >= 1_000_000:
                    profile["valuation_raw"] = raw
                    profile["valuation"] = val
                    break

        if not profile["valuation_raw"]:
            fin_result = ask_jamie(f"""Extract the financial size of {company_name} from this text.
Look for AUM, market cap, assets, valuation, funding.
Return JSON: {{"value": "10 billion", "type": "AUM"}} or {{"value": null}}
Text: {all_text[:2000]}
JSON:""", max_tokens=80, temperature=0.2)
            try:
                m = re.search(r'\{.*?\}', fin_result, re.DOTALL)
                if m:
                    d = json.loads(m.group())
                    if d.get("value"):
                        val = self.parse_value(str(d["value"]))
                        if val and val >= 1_000_000:
                            profile["valuation_raw"] = d["value"]
                            profile["valuation"] = val
            except: pass

        for kw, ctype in [
            ("family office","Family Office"),("sovereign wealth","Sovereign Wealth"),
            ("hedge fund","Investment Fund"),("private equity","Private Equity"),
            ("asset management","Asset Manager"),("endowment","Endowment"),
            ("nyse","Large Corp"),("nasdaq","Large Corp"),("lse","Large Corp"),
            ("series ","Funded Startup"),
        ]:
            if kw in all_text.lower():
                profile["company_type"] = ctype
                break

        qual = False
        qual_notes = []
        if profile["valuation"] and profile["valuation"] >= MIN_LIQUIDITY:
            qual = True
            qual_notes.append(f"✅ Value: {profile['valuation_raw']}")
        elif profile["company_type"] in ["Large Corp","Sovereign Wealth","Endowment"]:
            qual = True
            qual_notes.append(f"✅ {profile['company_type']} — auto-qualified")
        else:
            qual_notes.append("⚠️ Manual review needed")

        profile["qualified"] = qual
        profile["liquidity_flag"] = "QUALIFIED" if qual else "REVIEW NEEDED"
        profile["qualification_notes"] = qual_notes
        return profile

    def hunt_company(self, company_name, roles=None, pre_intel=None, account_context=None):
        """
        Hunt a company for executives.
        roles: list of roles to hunt — can be any role: CEO, CMO, CSR Director etc
        pre_intel: dict with location/aum/company_type already known from Phase 1
        account_context: dict with outreach_angle, email_tone etc from account profile
        """
        if roles is None:
            roles = account_context.get("target_roles", ["CEO","CFO","COO"]) if account_context else ["CEO","CFO","COO"]
        self.log = []
        self.think(f"🎯 HUNTING: {company_name}")

        domain = self.get_domain(company_name)
        profile = self.hunt_company_profile(company_name, domain, pre_intel=pre_intel)
        executives = []

        for role in roles:
            exec_data = self.hunt_executive(company_name, domain, role)
            exec_data["company"] = company_name
            exec_data["domain"] = domain
            executives.append(exec_data)
            time.sleep(0.5)

        found = len([e for e in executives if "[" not in (e["name"] or "")])
        self.think(f"✅ {company_name} | {profile['liquidity_flag']} | {found}/{len(roles)} found")

        return {"company":company_name,"domain":domain,"profile":profile,"executives":executives,"log":self.log,"hunt_date":datetime.now().isoformat()}

    def _normalise_name(self, name):
        """Lowercase, strip legal suffixes, remove non-alphanumeric — used for dedup"""
        n = re.sub(r'\s+(ltd|llc|inc|plc|limited|corp|corporation|group|holdings|co)\.?$', '', name, flags=re.IGNORECASE)
        return re.sub(r'[^a-z0-9]', '', n.lower().strip())

    def _upsert_company(self, company_name, domain, profile, account_id=None):
        """
        Find existing company by normalised name or domain, or create new one.
        Returns company_id (str) or None on failure.
        """
        headers = {"apikey":SUPABASE_KEY,"Authorization":f"Bearer {SUPABASE_KEY}",
                   "Content-Type":"application/json","Prefer":"return=representation"}
        norm = self._normalise_name(company_name)

        # 1. Match by normalised name + account_id
        try:
            params = f"name_normalised=eq.{norm}"
            if account_id:
                params += f"&account_id=eq.{account_id}"
            res = requests.get(f"{SUPABASE_URL}/rest/v1/companies?{params}&select=id",
                               headers={**headers, "Prefer":""}, timeout=6)
            if res.status_code == 200 and res.json():
                return res.json()[0]["id"]
        except Exception as e:
            print(f"  ⚠️ Company lookup error: {e}")

        # 2. Match by domain + account_id
        if domain:
            try:
                params = f"domain=eq.{domain}"
                if account_id:
                    params += f"&account_id=eq.{account_id}"
                res = requests.get(f"{SUPABASE_URL}/rest/v1/companies?{params}&select=id",
                                   headers={**headers, "Prefer":""}, timeout=6)
                if res.status_code == 200 and res.json():
                    return res.json()[0]["id"]
            except Exception as e:
                print(f"  ⚠️ Company domain lookup error: {e}")

        # 3. Insert new record
        aum_val = profile.get("valuation")
        record = {k:v for k,v in {
            "name":            company_name,
            "name_normalised": norm,
            "domain":          domain,
            "account_id":      account_id,
            "location":        profile.get("location"),
            "company_type":    profile.get("company_type"),
            "valuation":       profile.get("valuation_raw"),
            "aum_numeric":     aum_val if isinstance(aum_val, (int, float)) else None,
            "funding_status":  profile.get("funding_status"),
            "qualified":       profile.get("qualified", False),
            "liquidity_flag":  profile.get("liquidity_flag", "UNKNOWN"),
            "industry":        profile.get("industry"),
            "notes":           profile.get("notes"),
        }.items() if v is not None}

        try:
            res = requests.post(f"{SUPABASE_URL}/rest/v1/companies", headers=headers, json=record)
            if res.status_code in (200, 201) and res.json():
                cid = res.json()[0]["id"]
                print(f"  🏢 Company created: {company_name}")
                return cid
        except Exception as e:
            print(f"  ❌ Company insert error: {e}")
        return None

    def _save_community_sources(self, company_id, community_sources):
        """
        Save community sources (Discord/Telegram/Reddit/etc) linked to a company.
        Skips duplicates — checks url + company_id before inserting.
        """
        if not company_id or not community_sources:
            return
        headers = {"apikey":SUPABASE_KEY,"Authorization":f"Bearer {SUPABASE_KEY}",
                   "Content-Type":"application/json","Prefer":"return=representation"}
        for src in community_sources:
            url = src.get("url","")
            if not url:
                continue
            # Check for duplicate
            try:
                check = requests.get(
                    f"{SUPABASE_URL}/rest/v1/company_sources?company_id=eq.{company_id}&url=eq.{url}&select=id",
                    headers={**headers, "Prefer":""}, timeout=5)
                if check.status_code == 200 and check.json():
                    continue  # Already exists
            except:
                pass
            record = {k:v for k,v in {
                "company_id":  company_id,
                "source_type": src.get("source_type","web"),
                "name":        src.get("name",""),
                "url":         url,
                "notes":       src.get("notes",""),
            }.items() if v is not None}
            try:
                requests.post(f"{SUPABASE_URL}/rest/v1/company_sources",
                              headers=headers, json=record, timeout=5)
            except Exception as e:
                print(f"  ⚠️ Source save error: {e}")

    def push_to_crm(self, hunt_result, account_id=None, community_sources=None):
        headers = {"apikey":SUPABASE_KEY,"Authorization":f"Bearer {SUPABASE_KEY}","Content-Type":"application/json","Prefer":"return=representation"}
        added = []
        company = hunt_result["company"]
        profile = hunt_result.get("profile",{})

        # Upsert company record first — all contacts link to it
        company_id = self._upsert_company(company, hunt_result.get("domain"), profile, account_id)

        # Save community sources linked to this company
        if company_id:
            sources_to_save = community_sources or getattr(self.searcher, "_last_community_sources", [])
            self._save_community_sources(company_id, sources_to_save)

        for exec_data in hunt_result["executives"]:
            name = exec_data.get("name","")
            if not name or "[" in name: continue

            notes = "\n".join([
                f"── JAMIE Hunt: {datetime.now().strftime('%Y-%m-%d')} ──",
                f"Role: {exec_data['role']}",
                f"Confidence: {exec_data['confidence']}%",
                f"Method: {exec_data.get('method','Unknown')}",
                f"Brain: Abliterated Qwen3",
                f"",
                f"── Company Intel ──",
                f"Type: {profile.get('company_type','Unknown')}",
                f"HQ: {profile.get('location','Unknown')}",
                f"Valuation/AUM: {profile.get('valuation_raw','Unknown')}",
                f"Qualification: {profile.get('liquidity_flag','UNKNOWN')}",
                *profile.get("qualification_notes",[]),
                f"",
                f"LinkedIn: {exec_data.get('linkedin','Not found')}",
                *exec_data.get("notes",[]),
            ])

            contact = {k:v for k,v in {
                "name":name,"company":company,
                "email":exec_data.get("email"),
                "stage":"JAMIE HUNT","source":"JAMIE",
                "category":{"Family Office":"Investor","Investment Fund":"Investor","Sovereign Wealth":"Investor","Endowment":"Investor","Private Equity":"Investor","Asset Manager":"Investor"}.get(profile.get("company_type",""),"ARIA™ Client"),
                "notes":notes,"location":profile.get("location"),
                "valuation":profile.get("valuation_raw"),
                "liquidity_flag":profile.get("liquidity_flag","UNKNOWN"),
                "company_type":profile.get("company_type","Unknown"),
                "qualified":profile.get("qualified",False),
                "company_id":company_id,
                "account_id":account_id,
            }.items() if v is not None}

            try:
                res = requests.post(f"{SUPABASE_URL}/rest/v1/contacts",headers=headers,json=contact)
                if res.status_code in (200,201):
                    data = res.json()
                    contact_id = data[0]["id"] if data else None
                    added.append(name)
                    print(f"  ✅ CRM: {name} ({exec_data['role']}) — {'✅ QUALIFIED' if profile.get('qualified') else '⚠️ REVIEW'}")
                    if contact_id:
                        requests.post(f"{SUPABASE_URL}/rest/v1/communications",headers=headers,json={"contact_id":contact_id,"type":"Note","direction":"internal","notes":f"JAMIE Hunt {datetime.now().strftime('%Y-%m-%d %H:%M')}\nMethod: {exec_data.get('method')}\nConfidence: {exec_data['confidence']}%"})
            except Exception as e:
                print(f"  ❌ {e}")

        # Calculate JAMIE score and write back to company record
        if company_id:
            try:
                score_result = self.calculate_jamie_score(hunt_result)
                score        = score_result["total"]
                breakdown    = score_result["breakdown"]
                requests.patch(
                    f"{SUPABASE_URL}/rest/v1/companies?id=eq.{company_id}",
                    headers={**headers, "Prefer": ""},
                    json={"jamie_score": score},
                    timeout=6
                )
                print(f"  📊 JAMIE Score: {score}/100 ({score_result['label']}) — "
                      f"Names:{breakdown['name_confidence']} "
                      f"Email:{breakdown['email_confidence']} "
                      f"Profile:{breakdown['company_profile']} "
                      f"Finance:{breakdown['financial_qual']} "
                      f"Signals:{breakdown['engagement_signals']}")
                hunt_result["jamie_score"] = score_result
            except Exception as e:
                print(f"  ⚠️ Score write error: {e}")

        return added

    def calculate_jamie_score(self, hunt_result):
        """
        Score a company 0-100 across 5 dimensions after Phase 2 hunt.

        Breakdown (20pts each):
          name_confidence    — how many target roles were actually found
          email_confidence   — average email confidence across found executives
          company_profile    — how complete the company profile is
          financial_qual     — does AUM/valuation meet the $100m threshold
          engagement_signals — any comms logged (replies, opens etc)

        80-100 = Hot 🔥  |  60-79 = Warm ✓  |  40-59 = Cold  |  0-39 = Needs enrichment
        """
        profile    = hunt_result.get("profile", {})
        executives = hunt_result.get("executives", [])

        found_execs = [e for e in executives
                       if e.get("name") and "[" not in (e.get("name") or "")]
        total_roles = max(len(executives), 1)

        breakdown = {}

        # ── 1. Name confidence (0-20) ─────────────────────────
        # Full 20 if all roles found; partial credit for partial finds
        found_ratio = len(found_execs) / total_roles
        breakdown["name_confidence"] = round(found_ratio * 20)

        # ── 2. Email confidence (0-20) ────────────────────────
        # Average confidence score of found executives, scaled to 20
        if found_execs:
            avg_conf = sum(e.get("confidence", 0) for e in found_execs) / len(found_execs)
            breakdown["email_confidence"] = round((avg_conf / 100) * 20)
        else:
            breakdown["email_confidence"] = 0

        # ── 3. Company profile completeness (0-20) ────────────
        profile_fields = ["location", "company_type", "valuation_raw",
                          "liquidity_flag", "industry"]
        filled = sum(1 for f in profile_fields if profile.get(f) and
                     profile[f] not in ("Unknown", "UNKNOWN", None, ""))
        breakdown["company_profile"] = round((filled / len(profile_fields)) * 20)

        # ── 4. Financial qualification (0-20) ─────────────────
        val = profile.get("valuation")
        if val and isinstance(val, (int, float)) and val >= MIN_LIQUIDITY:
            breakdown["financial_qual"] = 20
        elif profile.get("company_type") in ["Sovereign Wealth", "Large Corp", "Endowment"]:
            breakdown["financial_qual"] = 20   # Auto-qualified types
        elif profile.get("qualified"):
            breakdown["financial_qual"] = 15   # Qualified but no exact figure
        elif val and isinstance(val, (int, float)) and val >= MIN_LIQUIDITY * 0.5:
            breakdown["financial_qual"] = 10   # Half threshold — partial credit
        else:
            breakdown["financial_qual"] = 0

        # ── 5. Engagement signals (0-20) ──────────────────────
        # At hunt time there are no engagement signals yet — starts at 0.
        # This dimension increases later as comms are logged (replied, opened).
        # We give 5pts just for having been hunted (data exists).
        breakdown["engagement_signals"] = 5

        total = sum(breakdown.values())
        total = max(0, min(100, total))  # Clamp 0-100

        if total >= 80:
            label = "Hot 🔥"
        elif total >= 60:
            label = "Warm ✓"
        elif total >= 40:
            label = "Cold"
        else:
            label = "Needs enrichment"

        return {"total": total, "breakdown": breakdown, "label": label}

    def get_strategy(self, hunt_result):
        profile = hunt_result.get("profile",{})
        execs = hunt_result.get("executives",[])
        exec_summary = "\n".join([f"- {e['role']}: {e['name']} ({e['confidence']}%, {e.get('email','no email')})" for e in execs if e.get("name") and "[" not in (e.get("name") or "")])
        return ask_groq(f"""ARIA analyst for ARIAtrust commodities.
Company: {hunt_result.get('company')} | Type: {profile.get('company_type')} | AUM: {profile.get('valuation_raw','Unknown')} | {profile.get('liquidity_flag')}
Executives: {exec_summary}
For each: outreach angle, specific hook, tone, send time. 3 sentences max per person.""")


# ── AUTONOMOUS HUNT ────────────────────────────────────────────

def autonomous_hunt(mission, count=50, roles=None):
    """
    Full autonomous hunt:
    1. Deep searcher finds company names
    2. JAMIE hunts each one individually
    3. Each gets pushed to CRM immediately
    """
    searcher = DeepSearcher()
    jamie    = JAMIEHunter()

    print(f"\n{'='*60}")
    print(f"  JAMIE — Autonomous Deep Hunt v5")
    print(f"  Brain: Abliterated Qwen3 (unrestricted)")
    print(f"  Mission: {mission}")
    print(f"  Target: {count} companies")
    print(f"{'='*60}\n")

    # Phase 1: Find targets
    print("PHASE 1: Finding targets...\n")
    companies = searcher.find_companies(mission, count)

    if not companies:
        print("  ❌ No companies found. Trying direct brain search...")
        # Last resort — ask brain directly
        result = ask_jamie(f"""List {min(count,30)} real company names that match: "{mission}"
These should be real organisations with $100m+ in assets.
Return ONLY a JSON array of company name strings.
Names:""", max_tokens=500, temperature=0.8)
        try:
            m = re.search(r'\[.*?\]', result, re.DOTALL)
            if m:
                companies = [n.strip() for n in json.loads(m.group()) if isinstance(n, str) and len(n.strip()) > 3]
                print(f"  Brain direct search found: {len(companies)} companies")
        except: pass

    if not companies:
        print("  ❌ Hunt failed. Try a more specific mission.")
        return []

    print(f"\n  Found {len(companies)} targets\n")
    print(f"PHASE 2: Hunting each company...\n")

    # Phase 2: Hunt each company one by one
    results = []
    qualified_count = 0
    total_added = 0

    for i, company in enumerate(companies):
        print(f"\n[{i+1}/{len(companies)}] {company}")
        print("-"*40)
        try:
            result = jamie.hunt_company(company, roles)
            added  = jamie.push_to_crm(result)
            if result.get("profile",{}).get("qualified"): qualified_count += 1
            total_added += len(added)
            result["crm_added"] = added
            results.append(result)
        except Exception as e:
            print(f"  ❌ Hunt failed: {e}")
        time.sleep(1.5)

    print(f"\n{'='*60}")
    print(f"  JAMIE Autonomous Hunt Complete!")
    print(f"  Companies found:    {len(companies)} (you asked for {count}, JAMIE found more!)") if len(companies) > count else print(f"  Companies found:    {len(companies)}")
    print(f"  Contacts added:     {total_added}")
    print(f"  Qualified ($100m+): {qualified_count}")
    print(f"{'='*60}\n")
    return results


def hunt_list(companies, roles=None):
    jamie = JAMIEHunter()
    results = []
    for i, company in enumerate(companies):
        print(f"\n[{i+1}/{len(companies)}] {company}")
        try:
            result = jamie.hunt_company(company, roles)
            added  = jamie.push_to_crm(result)
            result["crm_added"] = added
            results.append(result)
        except Exception as e:
            print(f"  ❌ {e}")
        time.sleep(1.5)
    return results


# ── CLI ────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("  JAMIE — Intelligence Hunting Engine v5")
    print("  Deep hunt | Multi-source | One by one")
    print("  Brain: Abliterated Qwen3 (unrestricted)")
    print("=" * 60)
    print("\n1. Single company hunt")
    print("2. Quick test (BlackRock)")
    print("3. Batch hunt (comma separated)")
    print("4. Autonomous deep hunt")
    print("5. Excel import hunt")

    choice = input("\nChoice: ").strip()

    if choice == "1":
        company = input("Company name: ").strip()
        jamie   = JAMIEHunter()
        result  = jamie.hunt_company(company)
        added   = jamie.push_to_crm(result)
        print(f"\nAdded: {added}")
        print(f"\nStrategy:\n{jamie.get_strategy(result)}")

    elif choice == "2":
        jamie  = JAMIEHunter()
        result = jamie.hunt_company("BlackRock", roles=["CEO","CFO"])
        added  = jamie.push_to_crm(result)
        print(f"\nAdded: {added}")
        print(f"\nStrategy:\n{jamie.get_strategy(result)}")

    elif choice == "3":
        raw = input("Companies (comma separated): ")
        companies = [c.strip() for c in raw.split(",") if c.strip()]
        hunt_list(companies)

    elif choice == "4":
        print("\nExamples:")
        print("  'family offices UK $100m AUM'")
        print("  'UAE investment funds $100m'")
        print("  'FTSE 250 companies'")
        print("  'US hedge funds $500m AUM'")
        mission = input("\nMission: ").strip()
        count   = int(input("Target count (default 20): ").strip() or "20")
        autonomous_hunt(mission, count)

    elif choice == "5":
        filepath = input("Excel file path: ").strip().strip('"')
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            companies = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and len(cell) > 2:
                        companies.append(cell.strip())
                        break
            print(f"\nFound {len(companies)} companies in Excel")
            confirm = input(f"Hunt all {len(companies)}? (y/n): ").strip().lower()
            if confirm == "y":
                hunt_list(companies)
        except Exception as e:
            print(f"Excel error: {e}")
